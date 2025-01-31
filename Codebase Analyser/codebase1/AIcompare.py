import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import openai
import time

highlighted_cells=set()

openai.api_key = "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"



def generate_response(text1,text2):
    openai.api_key = "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"
    response = openai.completions.create(
                    model="gpt-3.5-turbo-instruct",
                    prompt=f"Text 1: {text1}\nText 2: {text2}\nAre these two texts either value wise similar?\n",
                    max_tokens=1,
	)
    return response.choices[0].text.strip()


def compare_semantic_similarity(text1, text2):
    openai.api_key = "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"
    res=openai.similarity
    response = openai.completions.create(
        model="gpt-3.5-turbo-instruct",
        prompt=f"Text 1: {text1}\nText 2: {text2}\nAre these two texts either value wise similar?\n",
        max_tokens=1
    )
    return response.choices[0].text.strip()

def extract_value(value):
    """Extracts the numeric or alphabetic part from a string."""
    # Use regular expression to find all integers or alphabets in the string
    values = re.findall(r'\d+|[a-zA-Z]+', str(value))
    # If values are found, return the first one, else return None
    return values[0] if values else None

def compare_excel_sheets(file1, file2):
    # Load the Excel files
    wb1 = load_workbook(filename=file1)
    wb2 = load_workbook(filename=file2)

    # Select the first sheet from each workbook
    sheet1 = wb1.active
    sheet2 = wb2.active

    # Iterate over cells in both sheets
    for row in range(1, max(sheet1.max_row, sheet2.max_row) + 1):
        for col in range(1, max(sheet1.max_column, sheet2.max_column) + 1):
            # Get cell values from both sheets
            cell1 = sheet1.cell(row=row, column=col)
            cell2 = sheet2.cell(row=row, column=col)

            # Extract and compare values
            value1 = extract_value(cell1.value)
            value2 = extract_value(cell2.value)

            # Compare extracted values, ignoring capitalization
            if value1 and value2 and value1.lower() != value2.lower():
                # Set fill color to indicate difference
                cell1.fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")
                note = cell2.value
                cell1.comment = Comment(f"Value in Old Sheet: {note}", "Auto Generated")
                highlighted_cells.add(cell1.coordinate)
                # cell2.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Save the modified workbook
    print(highlighted_cells)
    i = 0
    for address in highlighted_cells:
        # Get cell values from both sheets
        cell1 = sheet1[address]
        cell2 = sheet2[address]
        if(i==3):
            time.sleep(60)
            i=0
        # Compare cell values using OpenAI's semantic similarity model
        similarity = compare_semantic_similarity(str(cell1.value), str(cell2.value))
        i=i+1
        print(similarity.lower())
        # Dehighlight cells if they are semantically similar
        if similarity.lower() == "yes":
            cell1.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    wb1.save(filename=file1)
    wb2.save(filename=file2)

def compare_excel_sheets_using_AI(file1, file2, cell_addresses):
    # Load the Excel files
    wb1 = load_workbook(filename=file1)
    wb2 = load_workbook(filename=file2)

    # Select the first sheet from each workbook
    sheet1 = wb1.active
    sheet2 = wb2.active

    # Iterate over specified cell addresses
    for address in cell_addresses:
        # Get cell values from both sheets
        cell1 = sheet1[address]
        cell2 = sheet2[address]

        # Compare cell values using OpenAI's semantic similarity model
        similarity = compare_semantic_similarity(str(cell1.value), str(cell2.value))
        print(similarity.lower())
        # Dehighlight cells if they are semantically similar
        if similarity.lower() == "yes":
            cell1.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
    wb1.save(filename=file1)
    wb2.save(filename=file2)

# Example usage
compare_excel_sheets("Competitor.xlsx", "Competitor_old_.xlsx")
#compare_excel_sheets_using_AI("Competitor.xlsx", "Competitor_old_.xlsx", highlighted_cells)