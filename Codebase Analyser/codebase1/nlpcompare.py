import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sentence_transformers import SentenceTransformer, util

def compare_and_format_excel(file_path1, sheet_name1, file_path2, sheet_name2, similarity_threshold=0.75):
    # Load Excel files
    df1 = pd.read_excel(file_path1, sheet_name=sheet_name1)
    df2 = pd.read_excel(file_path2, sheet_name=sheet_name2)

    # Load workbooks for formatting
    wb1 = load_workbook(file_path1)
    ws1 = wb1[sheet_name1]

    wb2 = load_workbook(file_path2)
    ws2 = wb2[sheet_name2]

    # Load the sentence transformer model
    model = SentenceTransformer('all-MiniLM-L6-v2')

    # Compare cells across sheets
    for row in range(1, min(df1.shape[0], df2.shape[0]) + 1):
        for col in range(1, df1.shape[1] + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)

            value1 = str(cell1.value)
            value2 = str(cell2.value)

            # Compute semantic similarity
            if value1 and value2:  # ensure both cells are not empty
                embeddings1 = model.encode(value1, convert_to_tensor=True)
                embeddings2 = model.encode(value2, convert_to_tensor=True)
                similarity = util.pytorch_cos_sim(embeddings1, embeddings2)

                if similarity >= similarity_threshold:
                    fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell1.fill = fill
                    cell2.fill = fill

    # Save the modified workbooks
    wb1.save("modified_" + file_path1)
    wb2.save("modified_" + file_path2)

# Example usage
compare_and_format_excel('Competitor.xlsx', 'Sheet1', 'Competitor_old_.xlsx', 'Sheet1')